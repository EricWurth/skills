#!/usr/bin/env python3
"""Generate templates/JobSearchTracker_Template.xlsx from scratch.

Three sheets: README (legend + rules), Tracker (headers, one fictional example
row, dropdown validation), Dashboard (live COUNTIF stats). Rerun this script to
regenerate the template after schema changes — never hand-edit the shipped file.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent.parent / "templates" / "JobSearchTracker_Template.xlsx"

COLUMNS = [
    ("id", 28, "kebab-case slug: company-short-title"),
    ("company", 20, "real employer (never the aggregator/agency)"),
    ("title", 32, "posted title"),
    ("status", 12, "new / ready / applied / interviewing / rejected / dead / deferred"),
    ("fit", 6, "1-5 score vs Profile/match-profile.md"),
    ("level", 18, "seniority read (judge by years-required, not title)"),
    ("effort", 10, "High / Med / Low-Med / Low prep tier"),
    ("variant", 16, "standing resume variant this role maps to"),
    ("comp", 16, "disclosed or estimated range"),
    ("compSource", 14, "posting / estimate / salary site"),
    ("location", 18, "as posted"),
    ("moveType", 14, "remote / local / relocation (+ exception tags)"),
    ("sector", 14, "industry"),
    ("found", 12, "date first seen YYYY-MM-DD"),
    ("board", 22, "where found, incl. the search that surfaced it"),
    ("url", 30, "posting URL on the board"),
    ("applyUrl", 30, "direct ATS apply URL (strip board redirects)"),
    ("ats", 14, "Workday / Greenhouse / Lever / iCIMS / custom"),
    ("notes", 40, "freeform; automation appends breadcrumbs"),
    ("packetComplete", 14, "TRUE only when the tailored resume file exists"),
    ("queueRank", 10, "apply order, 1 = first"),
    ("matchKey", 30, "lowercased company|title - THE dedupe key"),
    ("fitEvidence", 40, "one line: why the fit score"),
]

STATUSES = ["new", "ready", "applied", "interviewing", "rejected", "dead", "deferred"]
EFFORTS = ["High", "Med", "Low-Med", "Low"]

# One fictional example row (persona from examples/ExampleMasterResume.md).
EXAMPLE = [
    "acme-distribution-senior-ops-manager", "Acme Distribution (example)",
    "Senior Manager, Distribution Operations", "new", 4,
    "Senior Manager (8+ yrs required)", "Med", "Operations_Leader",
    "$115K-$140K", "posting", "Columbus, OH (hybrid)", "local", "logistics",
    "2026-01-15", "LinkedIn (keyword: distribution operations)",
    "https://example.com/job/123", "https://careers.example.com/job/123",
    "Workday", "Example row - replace with real finds", False, 1,
    "acme distribution (example)|senior manager, distribution operations",
    "Core WMS + team-size match; comp floor clears",
]

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
NOTE_FONT = Font(name=ARIAL, italic=True, color="595959", size=9)
DV_ROWS = 2000


def style_cell(c, bold=False, size=10, italic=False):
    c.font = Font(name=ARIAL, bold=bold, size=size, italic=italic)


def build_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95
    rows = [
        ("Job Search Tracker", "", 14, True),
        ("", "One row per job on the Tracker sheet. This sheet is the manual; delete it once you know the system.", 10, False),
        ("", "", 10, False),
        ("Fill in", "Everything on the Tracker sheet is yours to edit. Dropdowns: status, effort, packetComplete. The example row shows expected formats - overwrite it with your first real find.", 10, False),
        ("Dashboard", "The Dashboard sheet computes itself (live formulas). Don't type over it.", 10, False),
        ("", "", 10, False),
        ("Rule 1", "This file is the ONLY home for job-search state. No side lists.", 10, False),
        ("Rule 2", "matchKey = lowercase company|title. Same job on two boards is ONE row.", 10, False),
        ("Rule 3", "status becomes 'ready' only when packetComplete is TRUE and the tailored resume file actually exists.", 10, False),
        ("Rule 4", "Automation (resumebot skills) backs this file up to backups/ before every write and only appends or updates named fields.", 10, False),
        ("Rule 5", "Status lifecycle: new -> ready -> applied -> interviewing. Exits: rejected / dead / deferred (parked, not rejected).", 10, False),
        ("", "", 10, False),
        ("Columns", "", 11, True),
    ]
    r = 1
    for a, b, size, bold in rows:
        ca, cb = ws.cell(row=r, column=1, value=a or None), ws.cell(row=r, column=2, value=b or None)
        style_cell(ca, bold=bold, size=size)
        style_cell(cb, size=size)
        r += 1
    for name, _w, desc in COLUMNS:
        ca = ws.cell(row=r, column=1, value=name)
        cb = ws.cell(row=r, column=2, value=desc)
        style_cell(ca, bold=True)
        style_cell(cb)
        r += 1


def build_tracker(wb):
    ws = wb.create_sheet("Tracker")
    for i, (name, width, _desc) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    for i, val in enumerate(EXAMPLE, 1):
        c = ws.cell(row=2, column=i, value=val)
        c.font = NOTE_FONT

    dv_status = DataValidation(type="list", formula1=f'"{",".join(STATUSES)}"', allow_blank=True)
    dv_effort = DataValidation(type="list", formula1=f'"{",".join(EFFORTS)}"', allow_blank=True)
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_effort)
    ws.add_data_validation(dv_bool)
    dv_status.add(f"D2:D{DV_ROWS}")
    dv_effort.add(f"G2:G{DV_ROWS}")
    dv_bool.add(f"T2:T{DV_ROWS}")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    t = ws.cell(row=1, column=1, value="Pipeline Dashboard")
    style_cell(t, bold=True, size=13)
    note = ws.cell(row=2, column=1, value="Live formulas over the Tracker sheet - do not type here.")
    note.font = NOTE_FONT

    r = 4
    for s in STATUSES:
        style_cell(ws.cell(row=r, column=1, value=s), bold=(s in ("ready", "interviewing")))
        f = ws.cell(row=r, column=2, value=f'=COUNTIF(Tracker!$D$2:$D${DV_ROWS},"{s}")')
        style_cell(f)
        r += 1
    style_cell(ws.cell(row=r, column=1, value="total tracked"), bold=True)
    style_cell(ws.cell(row=r, column=2, value=f"=COUNTA(Tracker!$A$2:$A${DV_ROWS})"))
    r += 2
    style_cell(ws.cell(row=r, column=1, value="scored, packet pending"), bold=True)
    style_cell(ws.cell(row=r, column=2,
               value=f'=COUNTIFS(Tracker!$D$2:$D${DV_ROWS},"new",Tracker!$T$2:$T${DV_ROWS},FALSE)'))
    r += 1
    style_cell(ws.cell(row=r, column=1, value="ready to apply"), bold=True)
    style_cell(ws.cell(row=r, column=2,
               value=f'=COUNTIFS(Tracker!$D$2:$D${DV_ROWS},"ready",Tracker!$T$2:$T${DV_ROWS},TRUE)'))


def main():
    wb = openpyxl.Workbook()
    build_readme(wb)
    build_tracker(wb)
    build_dashboard(wb)
    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()

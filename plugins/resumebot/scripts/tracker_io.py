#!/usr/bin/env python3
"""Tracker I/O helpers for resumebot.

All automated access to JobSearchTracker.xlsx goes through this module so the
tracker rules (backup-before-write, fresh reads, matchKey dedupe, append-only,
backup retention) are enforced in one place.

CLI:
  python tracker_io.py filter  <xlsx> --where status=ready packetComplete=TRUE
  python tracker_io.py append  <xlsx> --rows rows.json --reason nightly_scan
  python tracker_io.py update  <xlsx> --id some-job-id --set status=applied --reason email_sync
  python tracker_io.py keys    <xlsx>            # all matchKeys (for cheap dedupe)
  python tracker_io.py counts  <xlsx>            # status counts

`rows.json` is a JSON list of objects keyed by column name. Unknown columns are
rejected. Rows whose matchKey already exists are skipped and reported.
"""

import argparse
import json
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SHEET = "Tracker"
BACKUP_KEEP = 10

COLUMNS = [
    "id", "company", "title", "status", "fit", "level", "effort", "variant",
    "comp", "compSource", "location", "moveType", "sector", "found", "board",
    "url", "applyUrl", "ats", "notes", "packetComplete", "queueRank",
    "matchKey", "fitEvidence",
]

STATUSES = {"new", "ready", "applied", "interviewing", "rejected", "dead", "deferred"}


def make_match_key(company: str, title: str) -> str:
    return f"{(company or '').strip().lower()}|{(title or '').strip().lower()}"


def backup(xlsx: Path, reason: str) -> Path:
    """Copy the tracker into Tracker/backups/ and prune old backups."""
    backups = xlsx.parent / "backups"
    backups.mkdir(exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_reason = "".join(c if c.isalnum() or c in "-_" else "_" for c in reason)[:40]
    dest = backups / f"{xlsx.stem}.{stamp}_{safe_reason}{xlsx.suffix}"
    shutil.copy2(xlsx, dest)
    old = sorted(backups.glob(f"{xlsx.stem}.*{xlsx.suffix}"))
    for stale in old[:-BACKUP_KEEP]:
        stale.unlink()
    return dest


def _open(xlsx: Path):
    wb = openpyxl.load_workbook(xlsx)
    if SHEET not in wb.sheetnames:
        sys.exit(f"error: sheet '{SHEET}' not found in {xlsx}")
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]
    if headers[: len(COLUMNS)] != COLUMNS:
        sys.exit(f"error: header mismatch in {xlsx}; expected {COLUMNS}")
    return wb, ws


def read_rows(xlsx: Path):
    wb, ws = _open(xlsx)
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(COLUMNS, r)))
    wb.close()
    return rows


def cmd_filter(args):
    rows = read_rows(Path(args.xlsx))
    conds = dict(kv.split("=", 1) for kv in args.where or [])
    out = []
    for row in rows:
        ok = True
        for col, want in conds.items():
            have = row.get(col)
            if isinstance(have, bool):
                have = "TRUE" if have else "FALSE"
            if str(have).strip().lower() != want.strip().lower():
                ok = False
                break
        if ok:
            out.append(row)
    json.dump(out, sys.stdout, indent=1, default=str)


def cmd_keys(args):
    keys = sorted({r["matchKey"] for r in read_rows(Path(args.xlsx)) if r.get("matchKey")})
    json.dump(keys, sys.stdout, indent=1)


def cmd_counts(args):
    counts = {}
    for r in read_rows(Path(args.xlsx)):
        s = str(r.get("status") or "?")
        counts[s] = counts.get(s, 0) + 1
    json.dump(counts, sys.stdout, indent=1)


def cmd_append(args):
    xlsx = Path(args.xlsx)
    new_rows = json.loads(Path(args.rows).read_text(encoding="utf-8"))
    for row in new_rows:
        unknown = set(row) - set(COLUMNS)
        if unknown:
            sys.exit(f"error: unknown columns {sorted(unknown)} in row {row.get('id')}")
        if row.get("status") and row["status"] not in STATUSES:
            sys.exit(f"error: bad status '{row['status']}' in row {row.get('id')}")
        if not row.get("matchKey"):
            row["matchKey"] = make_match_key(row.get("company", ""), row.get("title", ""))

    backup(xlsx, args.reason)
    wb, ws = _open(xlsx)  # fresh read AFTER backup, immediately before write
    existing = {str(r[COLUMNS.index("matchKey")]).lower()
                for r in ws.iter_rows(min_row=2, values_only=True)
                if r[COLUMNS.index("matchKey")]}
    added, skipped = [], []
    for row in new_rows:
        if row["matchKey"].lower() in existing:
            skipped.append(row["matchKey"])
            continue
        ws.append([row.get(c) for c in COLUMNS])
        existing.add(row["matchKey"].lower())
        added.append(row["matchKey"])
    wb.save(xlsx)
    json.dump({"added": added, "skipped_duplicates": skipped}, sys.stdout, indent=1)


def cmd_update(args):
    xlsx = Path(args.xlsx)
    updates = dict(kv.split("=", 1) for kv in args.set)
    unknown = set(updates) - set(COLUMNS)
    if unknown:
        sys.exit(f"error: unknown columns {sorted(unknown)}")
    if "status" in updates and updates["status"] not in STATUSES:
        sys.exit(f"error: bad status '{updates['status']}'")

    backup(xlsx, args.reason)
    wb, ws = _open(xlsx)
    id_col = COLUMNS.index("id") + 1
    hit = None
    for row in ws.iter_rows(min_row=2):
        if str(row[id_col - 1].value) == args.id:
            hit = row
            break
    if hit is None:
        sys.exit(f"error: id '{args.id}' not found")
    for col, val in updates.items():
        cell = hit[COLUMNS.index(col)]
        if col == "packetComplete":
            cell.value = str(val).strip().upper() == "TRUE"
        elif col == "notes" and args.append_note:
            cell.value = f"{cell.value or ''} | {val}".strip(" |")
        else:
            cell.value = val
    wb.save(xlsx)
    json.dump({"updated": args.id, "fields": updates}, sys.stdout, indent=1)


def main():
    p = argparse.ArgumentParser(description=__doc__)
    sub = p.add_subparsers(dest="cmd", required=True)

    f = sub.add_parser("filter"); f.add_argument("xlsx"); f.add_argument("--where", nargs="*"); f.set_defaults(fn=cmd_filter)
    k = sub.add_parser("keys"); k.add_argument("xlsx"); k.set_defaults(fn=cmd_keys)
    c = sub.add_parser("counts"); c.add_argument("xlsx"); c.set_defaults(fn=cmd_counts)

    a = sub.add_parser("append")
    a.add_argument("xlsx"); a.add_argument("--rows", required=True); a.add_argument("--reason", required=True)
    a.set_defaults(fn=cmd_append)

    u = sub.add_parser("update")
    u.add_argument("xlsx"); u.add_argument("--id", required=True); u.add_argument("--set", nargs="+", required=True)
    u.add_argument("--reason", required=True); u.add_argument("--append-note", action="store_true")
    u.set_defaults(fn=cmd_update)

    args = p.parse_args()
    args.fn(args)


if __name__ == "__main__":
    main()
